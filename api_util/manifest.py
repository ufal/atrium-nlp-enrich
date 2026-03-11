import os
import sys
import csv
import glob
import openpyxl



def get_sorted_text(file_path):
    """
    Reads a CSV or XLSX, sorts by page_num then line_num, and returns the full text.
    For XLSX, checks all sheets for a 'text' column.
    """
    entries = []
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.csv':
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        p = int(row.get('page_num', 0))
                    except (ValueError, TypeError):
                        p = 0
                    try:
                        l = int(row.get('line_num', 0))
                    except (ValueError, TypeError):
                        l = 0

                    text = row.get('text', '').strip()
                    if text:
                        entries.append({'p': p, 'l': l, 'text': text})
        except Exception as e:
            sys.stderr.write(f"Error parsing CSV {file_path}: {e}\n")

    elif ext == '.xlsx':
        if openpyxl is None:
            sys.stderr.write(
                "Error: 'openpyxl' module is required for .xlsx files. Install it using 'pip install openpyxl'\n")
            return None
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Extract headers from the first row
                headers = [cell.value for cell in sheet[1]]

                # Check if this sheet actually contains data and specifically the 'text' column
                if not headers or 'text' not in headers:
                    continue

                text_idx = headers.index('text')
                page_idx = headers.index('page_num') if 'page_num' in headers else -1
                line_idx = headers.index('line_num') if 'line_num' in headers else -1

                # Iterate over rows skipping the header
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    text_val = row[text_idx]
                    text = str(text_val).strip() if text_val is not None else ""

                    if text:
                        try:
                            p = int(row[page_idx]) if page_idx != -1 and row[page_idx] is not None else 0
                        except (ValueError, TypeError):
                            p = 0
                        try:
                            l = int(row[line_idx]) if line_idx != -1 and row[line_idx] is not None else 0
                        except (ValueError, TypeError):
                            l = 0

                        entries.append({'p': p, 'l': l, 'text': text})
        except Exception as e:
            sys.stderr.write(f"Error parsing XLSX {file_path}: {e}\n")
            return None

    else:
        sys.stderr.write(f"Unsupported file format: {file_path}\n")
        return None

    # Sort by Page (primary) and Line (secondary)
    entries.sort(key=lambda x: (x['p'], x['l']))

    return "\n".join([e['text'] for e in entries])



def main():
    INPUT_DIR = sys.argv[1]

    # Verify input directory
    if not os.path.isdir(INPUT_DIR):
        sys.stderr.write(f"Error: Directory {INPUT_DIR} not found.\n")
        sys.exit(1)

    # Find all CSV and XLSX files in the directory
    csv_files = glob.glob(os.path.join(INPUT_DIR, "*.csv"))
    xlsx_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))

    # Combine and sort all found files
    all_files = sorted(csv_files + xlsx_files)

    for file_path in all_files:
        # Extract Doc ID from filename (e.g., "CTX123.xlsx" -> "CTX123")
        filename = os.path.basename(file_path)
        doc_id = os.path.splitext(filename)[0]

        # Output format: DOC_ID [TAB] FILE_PATH
        print(f"{doc_id}\t{file_path}")


if __name__ == "__main__":
    main()

